from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def is_empty_row(values: list[Any]) -> bool:
    return all(v is None or str(v).strip() == "" for v in values)


def sheet_to_payload(ws) -> dict[str, Any]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"name": ws.title, "columns": [], "rows": []}

    headers_raw = list(rows[0])
    header_slots = []
    columns = []
    for idx, h in enumerate(headers_raw):
        name = normalize_header(h)
        if not name:
            continue
        header_slots.append((idx, name))
        columns.append(name)

    data_rows: list[dict[str, Any]] = []
    for row in rows[1:]:
        row_values = list(row)
        if is_empty_row(row_values):
            continue

        item: dict[str, Any] = {}
        has_value = False
        for idx, name in header_slots:
            value = row_values[idx] if idx < len(row_values) else None
            if value is None:
                item[name] = ""
            else:
                text = str(value)
                item[name] = text
                if text.strip() != "":
                    has_value = True

        if has_value:
            data_rows.append(item)

    return {
        "name": ws.title,
        "columns": columns,
        "rows": data_rows,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert workbook Excel to workbook JSON payload.")
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=False)
    args = parser.parse_args()

    input_path = Path(args.input)
    wb = load_workbook(input_path)
    payload = {"sheets": [sheet_to_payload(wb[sn]) for sn in wb.sheetnames]}

    output = Path(args.output) if args.output else input_path.with_suffix(".json")
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    total_rows = sum(len(s.get("rows", [])) for s in payload["sheets"])
    print(f"Converted {len(payload['sheets'])} sheets / {total_rows} rows -> {output}")


if __name__ == "__main__":
    main()
