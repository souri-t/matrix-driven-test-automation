from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook


def apply_column_widths(ws, headers: list[str], rows: list[list[Any]]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row in rows:
            value = "" if col_idx - 1 >= len(row) else row[col_idx - 1]
            max_len = max(max_len, len(str(value)) if value is not None else 0)
        letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 80)


def parse_rows(sheet_name: str, columns: list[str], rows_raw: Any) -> list[list[Any]]:
    if not isinstance(rows_raw, list):
        raise ValueError(f"Sheet '{sheet_name}': rows must be an array")

    rows: list[list[Any]] = []
    for i, row in enumerate(rows_raw, start=1):
        if isinstance(row, dict):
            rows.append([row.get(c, "") for c in columns])
        elif isinstance(row, list):
            rows.append(row)
        else:
            raise ValueError(
                f"Sheet '{sheet_name}' row {i}: each row must be an object or an array"
            )
    return rows


def write_workbook(root: dict[str, Any], output: Path) -> int:
    sheets = root.get("sheets")
    if not isinstance(sheets, list) or not sheets:
        raise ValueError("JSON root must include non-empty 'sheets' array")

    wb = Workbook()
    total_rows = 0
    first_sheet = True

    for i, sheet in enumerate(sheets, start=1):
        if not isinstance(sheet, dict):
            raise ValueError(f"Sheet index {i}: each sheet must be an object")

        name = sheet.get("name")
        columns = sheet.get("columns")
        rows_raw = sheet.get("rows", [])

        if not isinstance(name, str) or not name.strip():
            raise ValueError(f"Sheet index {i}: 'name' must be a non-empty string")
        if not isinstance(columns, list) or not columns or any(not isinstance(c, str) for c in columns):
            raise ValueError(f"Sheet '{name}': 'columns' must be a non-empty array of strings")

        rows = parse_rows(name, columns, rows_raw)

        if first_sheet:
            ws = wb.active
            ws.title = name
            first_sheet = False
        else:
            ws = wb.create_sheet(name)

        ws.append(columns)
        for row in rows:
            ws.append(row)

        apply_column_widths(ws, columns, rows)
        total_rows += len(rows)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return total_rows


def main() -> None:
    parser = argparse.ArgumentParser(description="Restore workbook Excel from workbook JSON payload.")
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    root = json.loads(Path(args.input).read_text(encoding="utf-8"))
    if not isinstance(root, dict):
        raise ValueError("JSON root must be an object")

    output = Path(args.output)
    count = write_workbook(root, output)
    print(f"Restored {count} rows -> {output}")


if __name__ == "__main__":
    main()
