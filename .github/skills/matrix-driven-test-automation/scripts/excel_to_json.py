from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook

from matrix_schema import REQUIRED_COLUMNS, normalize_column_name


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert matrix Excel to JSON.")
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    wb = load_workbook(args.input)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel is empty")

    header_raw = [str(c).strip() if c is not None else "" for c in rows[0]]
    headers = [normalize_column_name(c) for c in header_raw]

    missing = [c for c in REQUIRED_COLUMNS if c not in headers]
    if missing:
        raise ValueError(f"Missing required columns: {missing}")

    result: list[dict[str, str]] = []
    ids: set[str] = set()

    for row in rows[1:]:
        if row is None or all(v in (None, "") for v in row):
            continue

        item: dict[str, str] = {}
        for i, key in enumerate(headers):
            if not key:
                continue
            value = row[i] if i < len(row) else ""
            item[key] = "" if value is None else str(value).strip()

        rid = item.get("id", "")
        if not rid:
            raise ValueError("Found row with empty ID")
        if rid in ids:
            raise ValueError(f"Duplicated ID: {rid}")
        ids.add(rid)
        if not item.get("expected", ""):
            raise ValueError(f"Found row with empty expected (ID={rid})")

        result.append(item)

    output = Path(args.output)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"Converted {len(result)} cases -> {output}")


if __name__ == "__main__":
    main()
